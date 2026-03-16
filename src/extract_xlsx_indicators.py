#!/usr/bin/env python3
"""
Extract structured data from finance/tax xlsx/xls files for graph enrichment.

Reads Chinese-language tax spreadsheets and outputs clean JSON files with
standardized field names suitable for KuzuDB node creation.

Usage:
    python3 src/extract_xlsx_indicators.py                  # extract all 5 files
    python3 src/extract_xlsx_indicators.py --file <path>    # extract single file
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

# xlrd is required for .xls files
try:
    import xlrd
except ImportError:
    xlrd = None


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DOC_TAX_BASE = PROJECT_ROOT / "doc-tax" / "财税脑图3.0" / "1.财税脑图分类合集"
OUTPUT_DIR = PROJECT_ROOT / "data" / "extracted"

# --- File registry: (relative path from DOC_TAX_BASE, extractor function name) ---
FILE_REGISTRY = [
    ("税务稽查/各行业税负率及计算表.xlsx", "extract_industry_tax_burden"),
    ("税务稽查/2.66 税务预警指标测算系统2.09 (1).xlsx", "extract_tax_warning_indicators"),
    ("税务稽查/3.83 一般纳税人毛利率和税负关系实测模板1.5.xlsx", "extract_gross_margin_benchmarks"),
    ("发票相关/商品和服务税收分类编码表.xls", "extract_tax_classification_codes"),
    ("公司相关/3.40 纳税信用评价指标和评价方式.xls", "extract_tax_credit_indicators"),
]


def _parse_rate_line(text: str) -> dict | None:
    """Parse a line like '农副食品加工 3.50%' into (industry, rate)."""
    m = re.match(r"^(.+?)\s*([\d.]+)%$", text.strip())
    if m:
        return {"industry": m.group(1).strip(), "rate_pct": float(m.group(2))}
    # Handle lines without space before rate, e.g. '通信设备、计算机及其他电子设备制造业2.00%'
    m = re.match(r"^(.+?)([\d.]+)%$", text.strip())
    if m:
        return {"industry": m.group(1).strip(), "rate_pct": float(m.group(2))}
    return None


def extract_industry_tax_burden(filepath: Path) -> dict:
    """
    Extract industry-level VAT burden rates and income tax contribution rates.
    Sheet: '各行业税负率'
    Output: two lists — vat_burden_rates and income_tax_contribution_rates.
    These will become RiskIndicator nodes in the graph.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb["各行业税负率"]

    vat_rates = []
    income_tax_rates = []
    current_section = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        col_b = row[1].value if len(row) > 1 else None  # column B
        col_c = row[2].value if len(row) > 2 else None  # column C

        # Detect section headers
        if col_b and "增值税" in str(col_b) and "税负率" in str(col_b):
            current_section = "vat"
        if col_b and "企业" in str(col_b) and "所得税" in str(col_b) and "贡献率" in str(col_b):
            current_section = "income_tax"

        # Parse data lines from column C
        if col_c and current_section:
            parsed = _parse_rate_line(str(col_c))
            if parsed:
                entry = {
                    "industry": parsed["industry"],
                    "rate_pct": parsed["rate_pct"],
                }
                if current_section == "vat":
                    entry["indicator_type"] = "vat_burden_rate"
                    vat_rates.append(entry)
                else:
                    entry["indicator_type"] = "income_tax_contribution_rate"
                    income_tax_rates.append(entry)

    wb.close()
    return {
        "source_file": filepath.name,
        "node_type": "RiskIndicator",
        "vat_burden_rates": vat_rates,
        "income_tax_contribution_rates": income_tax_rates,
        "total_records": len(vat_rates) + len(income_tax_rates),
    }


def extract_tax_warning_indicators(filepath: Path) -> dict:
    """
    Extract tax warning (audit trigger) indicators from the '预警指标' sheet.
    Each row is a named indicator with formula, threshold, risk description,
    and audit focus areas. These become AuditTrigger nodes.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb["预警指标"]

    indicators = []
    for r in range(3, ws.max_row + 1):
        idx = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if idx is None or name is None:
            continue

        formula = ws.cell(r, 3).value or ""
        computed_value = ws.cell(r, 4).value
        threshold_raw = ws.cell(r, 5).value
        filter_result = ws.cell(r, 7).value  # check/cross mark
        risk_desc = ws.cell(r, 8).value or ""
        audit_focus = ws.cell(r, 9).value or ""

        # Normalize threshold
        threshold_value = None
        threshold_text = str(threshold_raw) if threshold_raw is not None else ""
        try:
            threshold_value = float(threshold_raw)
        except (TypeError, ValueError):
            pass

        indicators.append({
            "id": int(idx),
            "name": name.strip(),
            "formula": formula.strip().replace("\n", " "),
            "computed_value": computed_value,
            "threshold_value": threshold_value,
            "threshold_text": threshold_text.strip().replace("\n", " "),
            "triggered": filter_result in ("√", "✓", True),
            "risk_description": risk_desc.strip().replace("\n", " "),
            "audit_focus": audit_focus.strip().replace("\n", " "),
        })

    wb.close()
    return {
        "source_file": filepath.name,
        "node_type": "AuditTrigger",
        "indicators": indicators,
        "total_records": len(indicators),
    }


def extract_gross_margin_benchmarks(filepath: Path) -> dict:
    """
    Extract gross margin vs tax burden relationship data from two sheets:
    - '经典案例': textbook example of margin-to-tax-burden mapping
    - '实测模版': real-world testing template with line items

    These become benchmark reference data for RiskIndicator enrichment.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    # Sheet 1: classic case
    ws1 = wb["经典案例"]
    classic_case = {}
    for r in range(3, ws1.max_row + 1):
        item = ws1.cell(r, 2).value
        amount = ws1.cell(r, 3).value
        formula = ws1.cell(r, 4).value or ""
        remark = ws1.cell(r, 5).value or ""
        if item:
            classic_case[item.strip()] = {
                "amount": amount,
                "formula": formula.strip(),
                "remark": remark.strip().replace("\n", " "),
            }

    # Sheet 2: testing template
    ws2 = wb["实测模版"]
    template_items = []
    for r in range(3, ws2.max_row + 1):
        idx = ws2.cell(r, 1).value
        item = ws2.cell(r, 3).value
        amount = ws2.cell(r, 4).value
        source = ws2.cell(r, 5).value or ""
        remark = ws2.cell(r, 6).value or ""
        if item:
            template_items.append({
                "seq": idx,
                "item": item.strip(),
                "amount": amount,
                "data_source": source.strip(),
                "remark": remark.strip().replace("\n", " "),
            })

    wb.close()
    return {
        "source_file": filepath.name,
        "node_type": "GrossMarginBenchmark",
        "classic_case": classic_case,
        "testing_template": template_items,
        "total_records": len(classic_case) + len(template_items),
    }


def extract_tax_classification_codes(filepath: Path) -> dict:
    """
    Extract the national tax classification code table (商品和服务税收分类编码表).
    Sheet: '明细表', ~4200 rows of hierarchical codes.
    Critical for invoice-to-product mapping in the graph.
    """
    if xlrd is None:
        raise ImportError("xlrd is required for .xls files: pip install xlrd")

    wb = xlrd.open_workbook(str(filepath))
    ws = wb.sheet_by_name("明细表")

    # Row 0: headers [序号, 编码(篇/类/章/节/条/款/项/目/子目/细目), 合并编码, 货物和劳务名称, 商品和服务分类简称, 说明]
    # Row 1: sub-headers for code segments
    # Data starts at row 2

    codes = []
    for r in range(2, ws.nrows):
        row_vals = [ws.cell_value(r, c) for c in range(ws.ncols)]

        seq = row_vals[0]  # ordinal
        # Code segments: cols 1-10 (篇,类,章,节,条,款,项,目,子目,细目)
        code_segments = [str(row_vals[c]).strip() for c in range(1, 11)]
        combined_code = str(row_vals[11]).strip() if len(row_vals) > 11 else ""
        item_name = str(row_vals[12]).strip() if len(row_vals) > 12 else ""
        category_abbr = str(row_vals[13]).strip() if len(row_vals) > 13 else ""
        description = str(row_vals[14]).strip() if len(row_vals) > 14 else ""

        # Skip empty rows
        if not item_name or item_name == "":
            continue

        # Determine hierarchy level by counting non-zero segments
        level = 0
        for seg in code_segments:
            if seg and seg != "00" and seg != "0" and seg != "":
                level += 1

        codes.append({
            "seq": _to_int_safe(seq),
            "code": combined_code,
            "code_segments": code_segments,
            "level": level,
            "item_name": item_name,
            "category_abbr": category_abbr,
            "description": description if description else None,
        })

    return {
        "source_file": filepath.name,
        "node_type": "TaxClassificationCode",
        "codes": codes,
        "total_records": len(codes),
    }


def extract_tax_credit_indicators(filepath: Path) -> dict:
    """
    Extract taxpayer credit evaluation indicators from the '指标' sheet.
    Hierarchical structure: category > level1 > level2 > level3 indicators,
    each with deduction points or direct grade assignment.
    """
    if xlrd is None:
        raise ImportError("xlrd is required for .xls files: pip install xlrd")

    wb = xlrd.open_workbook(str(filepath))
    ws = wb.sheet_by_name("指标")

    indicators = []
    current_info_source = ""  # '税务内部信息' or '外部信息'
    current_freq = ""         # '经常性指标信息' or '非经常性指标信息'
    current_l1 = ""
    current_l2 = ""

    for r in range(10, ws.nrows):
        vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
        non_empty = [v for v in vals if v != ""]
        if not non_empty:
            continue

        # Track hierarchical context from merged cells
        if vals[0] and "信息" in str(vals[0]):
            current_info_source = str(vals[0]).strip()
        if vals[1] and "指标信息" in str(vals[1]):
            current_freq = str(vals[1]).strip()

        # Detect level-1 indicator codes (e.g. "01.涉税申报信息")
        for i, v in enumerate(vals):
            v_str = str(v).strip()
            if re.match(r"^0\d\.", v_str) and len(v_str) > 3:
                if i <= 2:
                    current_l1 = v_str
                    break

        # Detect level-2 indicator codes (e.g. "0101.按照规定申报纳税")
        for i, v in enumerate(vals):
            v_str = str(v).strip()
            if re.match(r"^0\d0\d\.", v_str) and len(v_str) > 5:
                current_l2 = v_str
                break

        # Extract level-3 indicators with deduction points
        # They have codes like "010101." and a deduction score
        for i, v in enumerate(vals):
            v_str = str(v).strip()
            if not re.match(r"^0\d0\d0\d", v_str):
                continue

            # Found a level-3 indicator
            indicator_text = v_str
            deduction = ""
            direct_grade = ""

            # Scan remaining columns for deduction or grade
            for j in range(i + 1, len(vals)):
                cell = str(vals[j]).strip()
                if "分" in cell and cell != "":
                    deduction = cell
                if "直接判" in cell or "直接  判" in cell:
                    direct_grade = cell.replace("  ", "")

            # Extract code prefix
            code_match = re.match(r"^(\d{6})\.", indicator_text)
            code = code_match.group(1) if code_match else ""
            desc = indicator_text[len(code) + 1:].strip() if code else indicator_text

            # Handle compound indicators (multiple codes in one cell, separated by ;)
            sub_indicators = re.split(r";\s*(?=0\d0\d0\d)", indicator_text)
            if len(sub_indicators) > 1:
                for sub in sub_indicators:
                    sm = re.match(r"^(\d{6})\.(.*)", sub.strip())
                    if sm:
                        indicators.append({
                            "code": sm.group(1),
                            "description": sm.group(2).strip(),
                            "info_source": current_info_source,
                            "frequency": current_freq,
                            "level1": current_l1,
                            "level2": current_l2,
                            "deduction_points": deduction,
                            "direct_grade": direct_grade if direct_grade else None,
                        })
            else:
                indicators.append({
                    "code": code,
                    "description": desc,
                    "info_source": current_info_source,
                    "frequency": current_freq,
                    "level1": current_l1,
                    "level2": current_l2,
                    "deduction_points": deduction,
                    "direct_grade": direct_grade if direct_grade else None,
                })
            break  # only process first level-3 match per row

    # Also extract external evaluation info (rows 72+)
    external_indicators = []
    for r in range(72, ws.nrows):
        vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
        non_empty = [str(v).strip() for v in vals if v != ""]
        if len(non_empty) >= 2:
            # External rows have: category, description, scoring
            external_indicators.append({
                "category": non_empty[0] if len(non_empty) > 0 else "",
                "description": non_empty[1] if len(non_empty) > 1 else "",
                "scoring": non_empty[2] if len(non_empty) > 2 else "",
            })

    return {
        "source_file": filepath.name,
        "node_type": "TaxCreditIndicator",
        "internal_indicators": indicators,
        "external_indicators": external_indicators,
        "total_records": len(indicators) + len(external_indicators),
    }


def _to_int_safe(val) -> int | None:
    """Safely convert a value to int."""
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def write_json(data: dict, output_name: str) -> Path:
    """Write extracted data to JSON file."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / output_name
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    return out_path


def run_all():
    """Run all extractors and write output files."""
    extractors = {
        "extract_industry_tax_burden": (extract_industry_tax_burden, "industry_tax_burden_rates.json"),
        "extract_tax_warning_indicators": (extract_tax_warning_indicators, "tax_warning_indicators.json"),
        "extract_gross_margin_benchmarks": (extract_gross_margin_benchmarks, "gross_margin_benchmarks.json"),
        "extract_tax_classification_codes": (extract_tax_classification_codes, "tax_classification_codes.json"),
        "extract_tax_credit_indicators": (extract_tax_credit_indicators, "tax_credit_indicators.json"),
    }

    results = []
    for rel_path, func_name in FILE_REGISTRY:
        filepath = DOC_TAX_BASE / rel_path
        if not filepath.exists():
            print(f"WARN: file not found: {filepath}", file=sys.stderr)
            continue

        func, output_name = extractors[func_name]
        print(f"Extracting: {filepath.name} -> {output_name}")
        try:
            data = func(filepath)
            out_path = write_json(data, output_name)
            print(f"  OK: {data['total_records']} records -> {out_path}")
            results.append({"file": output_name, "records": data["total_records"], "status": "OK"})
        except Exception as e:
            print(f"  ERROR: {e}", file=sys.stderr)
            results.append({"file": output_name, "records": 0, "status": f"ERROR: {e}"})

    # Summary
    print(f"\n--- Summary: {len(results)} files processed ---")
    for r in results:
        print(f"  {r['status']:>5}  {r['records']:>5} records  {r['file']}")

    return results


def run_single(filepath_str: str):
    """Run extraction on a single file (auto-detect extractor by filename)."""
    filepath = Path(filepath_str).resolve()
    if not filepath.exists():
        print(f"ERROR: file not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    # Match by filename substring
    name = filepath.name
    extractors_map = {
        "各行业税负率": (extract_industry_tax_burden, "industry_tax_burden_rates.json"),
        "税务预警指标": (extract_tax_warning_indicators, "tax_warning_indicators.json"),
        "毛利率和税负": (extract_gross_margin_benchmarks, "gross_margin_benchmarks.json"),
        "税收分类编码": (extract_tax_classification_codes, "tax_classification_codes.json"),
        "纳税信用评价": (extract_tax_credit_indicators, "tax_credit_indicators.json"),
    }

    for key, (func, output_name) in extractors_map.items():
        if key in name:
            print(f"Extracting: {name} -> {output_name}")
            data = func(filepath)
            out_path = write_json(data, output_name)
            print(f"  OK: {data['total_records']} records -> {out_path}")
            return

    print(f"ERROR: no extractor matches filename: {name}", file=sys.stderr)
    sys.exit(1)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract structured data from tax xlsx/xls files")
    parser.add_argument("--file", "-f", help="Single file to extract (auto-detect extractor)")
    args = parser.parse_args()

    if args.file:
        run_single(args.file)
    else:
        run_all()
