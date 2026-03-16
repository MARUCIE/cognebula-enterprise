#!/usr/bin/env python3
"""P0-3: Extract top-level xlsx files and compliance checklist PDF.

Files:
- 财智云慧 0123-云苍穹财务分析参考.xlsx
- 财智云慧 0123-云苍穹数据源参考.xlsx
- 智能经营分析结构化思考V4@1028.xlsx
- 福建省民营企业涉税合规自查手册.pdf

Output: data/extracted/business_analysis/
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl

BASE_DIR = Path("/Users/mauricewen/Projects/cognebula-enterprise")
OUTPUT_DIR = BASE_DIR / "data" / "extracted" / "business_analysis"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

XLSX_FILES = [
    BASE_DIR / "doc-tax" / "财智云慧 0123-云苍穹财务分析参考.xlsx",
    BASE_DIR / "doc-tax" / "财智云慧 0123-云苍穹数据源参考.xlsx",
    BASE_DIR / "doc-tax" / "智能经营分析结构化思考V4@1028.xlsx",
]

PDF_FILE = BASE_DIR / "doc-tax" / "福建省民营企业涉税合规自查手册.pdf"


def extract_xlsx(xlsx_path: Path) -> dict:
    """Extract sheet names, headers, and data from an xlsx file."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    result = {
        "file": xlsx_path.name,
        "file_size_bytes": xlsx_path.stat().st_size,
        "sheet_count": len(wb.sheetnames),
        "sheets": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])

        if not rows:
            result["sheets"].append({
                "name": sheet_name,
                "headers": [],
                "row_count": 0,
                "sample_rows": [],
            })
            continue

        # First non-empty row as headers
        headers = rows[0] if rows else []
        data_rows = rows[1:]

        # Filter out completely empty rows
        data_rows = [r for r in data_rows if any(c.strip() for c in r)]

        sheet_data = {
            "name": sheet_name,
            "headers": headers,
            "row_count": len(data_rows),
            "column_count": len(headers),
            "sample_rows": data_rows[:5],  # first 5 data rows as sample
            "data": data_rows,  # full data for knowledge extraction
        }
        result["sheets"].append(sheet_data)

    wb.close()
    return result


def extract_compliance_pdf(pdf_path: Path) -> dict:
    """Extract checklist structure from the compliance self-check handbook.

    This PDF is scanned (image-based, 0 text chars per page).
    Strategy: extract page images, record structure metadata.
    If text extraction yields nothing, note it as scanned and extract
    whatever structure is available (TOC from bookmarks, image count).
    """
    doc = fitz.open(str(pdf_path))
    result = {
        "file": pdf_path.name,
        "file_size_bytes": pdf_path.stat().st_size,
        "page_count": len(doc),
        "chapters": [],
        "checklist_items": [],
        "tables": [],
        "is_scanned": False,
    }

    # Check if this is a scanned PDF (no text on first 5 pages)
    text_chars = 0
    for i in range(min(5, len(doc))):
        text_chars += len(doc[i].get_text("text").strip())

    if text_chars == 0:
        result["is_scanned"] = True
        print("(scanned PDF, no embedded text) ", end="", flush=True)

        # Try to extract TOC/bookmarks
        toc = doc.get_toc()
        if toc:
            result["chapters"] = [
                {"title": entry[1], "page": entry[2], "level": entry[0]}
                for entry in toc
            ]

        # Record page image info
        page_info = []
        for i in range(len(doc)):
            page = doc[i]
            images = page.get_images()
            page_info.append({
                "page": i + 1,
                "image_count": len(images),
                "width": round(page.rect.width),
                "height": round(page.rect.height),
            })
        result["page_images"] = page_info
        result["metadata"] = {
            "total_chapters": len(result["chapters"]),
            "total_sections": 0,
            "total_checklist_items": 0,
            "total_tables": 0,
            "note": "Scanned PDF - requires OCR (tesseract chi_sim) for text extraction",
        }
        doc.close()
        return result

    # Normal text-based PDF extraction
    current_chapter = None
    current_section = None

    chapter_re = re.compile(r"^[一二三四五六七八九十]+[、.．]|^第[一二三四五六七八九十]+[章节篇部分]")
    section_re = re.compile(r"^[（(]\s*[一二三四五六七八九十\d]+\s*[）)]|^\d+\.\d+")
    checklist_re = re.compile(r"[□√×✓✗☐☑]|自查|检查|核查|是否|合规|违规|风险点")

    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text("text")

        for line in text.split("\n"):
            stripped = line.strip()
            if not stripped:
                continue

            if chapter_re.match(stripped):
                current_chapter = {
                    "title": stripped,
                    "page": page_num + 1,
                    "sections": [],
                }
                result["chapters"].append(current_chapter)
                current_section = None
            elif section_re.match(stripped):
                current_section = {
                    "title": stripped,
                    "page": page_num + 1,
                }
                if current_chapter:
                    current_chapter["sections"].append(current_section)

            if checklist_re.search(stripped):
                result["checklist_items"].append({
                    "text": stripped,
                    "page": page_num + 1,
                    "chapter": current_chapter["title"] if current_chapter else None,
                })

        # Extract tables
        tables = page.find_tables()
        for table in tables:
            extracted = table.extract()
            if extracted and len(extracted) > 1:
                headers = [str(c).strip() if c else "" for c in extracted[0]]
                data_rows = []
                for row in extracted[1:]:
                    data_rows.append([str(c).strip() if c else "" for c in row])
                result["tables"].append({
                    "page": page_num + 1,
                    "headers": headers,
                    "row_count": len(data_rows),
                    "data": data_rows,
                    "chapter": current_chapter["title"] if current_chapter else None,
                })

    doc.close()

    result["metadata"] = {
        "total_chapters": len(result["chapters"]),
        "total_sections": sum(len(c.get("sections", [])) for c in result["chapters"]),
        "total_checklist_items": len(result["checklist_items"]),
        "total_tables": len(result["tables"]),
    }
    return result


def main():
    manifest = {
        "extracted_at": datetime.now().isoformat(),
        "output_dir": str(OUTPUT_DIR),
        "files": [],
    }

    # Extract xlsx files
    for xlsx_path in XLSX_FILES:
        if not xlsx_path.exists():
            print(f"  WARN: File not found: {xlsx_path.name}")
            manifest["files"].append({"file": xlsx_path.name, "error": "not found"})
            continue

        print(f"  Extracting: {xlsx_path.name} ...", end=" ", flush=True)
        try:
            result = extract_xlsx(xlsx_path)
            slug = xlsx_path.stem.replace(" ", "_")
            out_path = OUTPUT_DIR / f"{slug}.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(result, f, ensure_ascii=False, indent=2)

            total_rows = sum(s["row_count"] for s in result["sheets"])
            print(f"OK ({result['sheet_count']} sheets, {total_rows} rows)")
            manifest["files"].append({
                "file": xlsx_path.name,
                "output": out_path.name,
                "sheet_count": result["sheet_count"],
                "total_rows": total_rows,
            })
        except Exception as e:
            print(f"ERROR: {e}")
            manifest["files"].append({"file": xlsx_path.name, "error": str(e)})

    # Extract compliance PDF
    if PDF_FILE.exists():
        print(f"  Extracting: {PDF_FILE.name} ...", end=" ", flush=True)
        try:
            result = extract_compliance_pdf(PDF_FILE)
            out_path = OUTPUT_DIR / "fujian_compliance_checklist.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(result, f, ensure_ascii=False, indent=2)

            meta = result["metadata"]
            print(f"OK ({meta['total_chapters']} chapters, {meta['total_checklist_items']} checklist items, "
                  f"{meta['total_tables']} tables)")
            manifest["files"].append({
                "file": PDF_FILE.name,
                "output": out_path.name,
                **meta,
            })
        except Exception as e:
            print(f"ERROR: {e}")
            manifest["files"].append({"file": PDF_FILE.name, "error": str(e)})
    else:
        print(f"  WARN: File not found: {PDF_FILE.name}")

    # Write manifest
    manifest_path = OUTPUT_DIR / "manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    print(f"\nP0-3 DONE: {len(manifest['files'])} files processed")
    print(f"  Output: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
