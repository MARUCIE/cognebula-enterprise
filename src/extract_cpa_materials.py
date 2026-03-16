#!/usr/bin/env python3
"""Extract structured content from CPA study PDFs + XLSX for knowledge graph enrichment.

Reads CPA exam prep materials (journal entries, formulas, tax rules, audit points,
lecture notes, practice problems), classifies content by type, and outputs JSON
files to data/extracted/cpa/ with a manifest.

Usage:
    .venv/bin/python3 src/extract_cpa_materials.py
    .venv/bin/python3 src/extract_cpa_materials.py --dry-run
"""

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import fitz  # PyMuPDF
except ImportError:
    sys.exit("ERROR: PyMuPDF not installed. Run: pip install PyMuPDF")

try:
    import openpyxl
except ImportError:
    openpyxl = None  # graceful degradation; only needed for xlsx

# ── Constants ────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DOC_BASE = PROJECT_ROOT / "doc-tax" / "09CPA学习"
OUTPUT_DIR = PROJECT_ROOT / "data" / "extracted" / "cpa"

# Content type classification patterns
RE_JOURNAL_ENTRY = re.compile(r"[借贷][:：]")
RE_FORMULA = re.compile(r"[＝=×÷%‰∑√∫≥≤±]+")
RE_TAX_RATE = re.compile(r"(?:税率|征收率|扣除率|退税率|免征|减按|加计)\s*[:：]?\s*\d")
RE_HEADING = re.compile(
    r"^(?:第[一二三四五六七八九十百千\d]+[章节篇]|"
    r"[一二三四五六七八九十]+[、.]\s*|"
    r"\d{1,3}[、.\s]|"
    r"[（(]\s*[一二三四五六七八九十\d]+\s*[）)])",
    re.MULTILINE,
)
RE_CHOICE_Q = re.compile(r"[（(]\s*[ABCD]\s*[）)]")


# ── File Registry ────────────────────────────────────────────────────────

# Priority files with explicit classification hints
PRIORITY_FILES = [
    {
        "rel_path": "2021版高顿CPA内部资料（会计审计税法财管）/CPA会计加油包/02-《CPA会计-搞定会计必会经典分录》.pdf",
        "label": "classic_journal_entries",
        "subject": "accounting",
        "content_hint": "journal_entries",
    },
    {
        "rel_path": "2021版高顿CPA内部资料（会计审计税法财管）/CPA会计加油包/01-《CPA会计-高频考点集锦之考点妙记》.pdf",
        "label": "accounting_key_points",
        "subject": "accounting",
        "content_hint": "key_points",
    },
    {
        "rel_path": "2021版高顿CPA内部资料（会计审计税法财管）/CPA税法加油包/01-《CPA考点妙记-税法》.pdf",
        "label": "tax_law_key_points",
        "subject": "tax_law",
        "content_hint": "tax_rules",
    },
    {
        "rel_path": "2021版高顿CPA内部资料（会计审计税法财管）/CPA税法加油包/02-《CPA税法-不可忽视的税收政策》.pdf",
        "label": "important_tax_policies",
        "subject": "tax_law",
        "content_hint": "tax_rules",
    },
    {
        "rel_path": "2021版高顿CPA内部资料（会计审计税法财管）/CPA审计加油包/02-《CPA审计-易错易混知识点归纳》.pdf",
        "label": "audit_confusion_points",
        "subject": "audit",
        "content_hint": "audit_points",
    },
    {
        "rel_path": "2021版高顿CPA内部资料（会计审计税法财管）/CPA审计加油包/01-《CPA审计-高频考点集锦之考点妙记》.pdf",
        "label": "audit_key_points",
        "subject": "audit",
        "content_hint": "audit_points",
    },
    {
        "rel_path": "2021版高顿CPA内部资料（会计审计税法财管）/CPA审计加油包/03-审计-命题趋势分析与备考建议.pdf",
        "label": "audit_exam_trends",
        "subject": "audit",
        "content_hint": "audit_points",
    },
    {
        "rel_path": "2021版高顿CPA内部资料（会计审计税法财管）/CPA财管加油包/01-《CPA财管-高频考点集锦之考点妙记》.pdf",
        "label": "financial_mgmt_key_points",
        "subject": "financial_management",
        "content_hint": "formulas",
    },
    {
        "rel_path": "2021版高顿CPA内部资料（会计审计税法财管）/CPA财管加油包/02-《CPA财管-拿下财管必记公式图表》.pdf",
        "label": "financial_mgmt_formulas",
        "subject": "financial_management",
        "content_hint": "formulas",
    },
    {
        "rel_path": "CPA财管公式大全（一图汇总全部公式）/CPA财管公式大全（一图汇总全部公式）/注册会计师《财务成本管理》公式大全.pdf",
        "label": "cost_mgmt_formula_collection",
        "subject": "financial_management",
        "content_hint": "formulas",
    },
    {
        "rel_path": "CPA财管公式大全（一图汇总全部公式）/CPA财管公式大全（一图汇总全部公式）/CPA 会计常用 50 个公式！.pdf",
        "label": "top_50_accounting_formulas",
        "subject": "accounting",
        "content_hint": "formulas",
    },
]

# Directory-based batch collections
BATCH_DIRS = [
    {
        "rel_dir": "02逆袭精讲班-会计",
        "label_prefix": "lecture_accounting",
        "subject": "accounting",
        "content_hint": "lecture_notes",
    },
    {
        "rel_dir": "03进阶习题班-会计",
        "label_prefix": "practice_accounting",
        "subject": "accounting",
        "content_hint": "practice_problems",
    },
    {
        "rel_dir": "02逆袭精讲班-财管",
        "label_prefix": "lecture_financial_mgmt",
        "subject": "financial_management",
        "content_hint": "lecture_notes",
    },
    {
        "rel_dir": "03进阶习题班-财管",
        "label_prefix": "practice_financial_mgmt",
        "subject": "financial_management",
        "content_hint": "practice_problems",
    },
]

# XLSX files
XLSX_FILES = [
    {
        "rel_path": "会计经典分录汇总/会计经典分录汇总.xlsx",
        "label": "classic_journal_entries_xlsx",
        "subject": "accounting",
        "content_hint": "journal_entries",
    },
]


# ── Extraction helpers ───────────────────────────────────────────────────


def file_hash(path: Path) -> str:
    """SHA-256 of first 64KB (fast fingerprint)."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        h.update(f.read(65536))
    return h.hexdigest()[:16]


def extract_pdf_text(pdf_path: Path) -> list[dict]:
    """Extract text from each page of a PDF. Returns list of {page, text}."""
    pages = []
    try:
        doc = fitz.open(str(pdf_path))
        for i, page in enumerate(doc):
            text = page.get_text().strip()
            if text and len(text) > 10:  # skip near-empty pages (cover, blank)
                pages.append({"page": i + 1, "text": text})
        doc.close()
    except Exception as e:
        print(f"  WARN: failed to read {pdf_path.name}: {e}")
    return pages


def extract_xlsx_text(xlsx_path: Path) -> list[dict]:
    """Extract rows from XLSX as structured text. Returns list of {sheet, rows}."""
    if openpyxl is None:
        print("  WARN: openpyxl not installed, skipping XLSX")
        return []
    sheets = []
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                line = " | ".join(c for c in cells if c)
                if line:
                    rows.append(line)
            if rows:
                sheets.append({"sheet": name, "rows": rows})
        wb.close()
    except Exception as e:
        print(f"  WARN: failed to read {xlsx_path.name}: {e}")
    return sheets


def classify_page(text: str, hint: str) -> list[str]:
    """Classify a page's content types based on pattern matching + hint.

    Returns a list of detected content types. The hint provides a default
    classification from the file metadata.
    """
    types = set()

    # Pattern-based detection
    journal_count = len(RE_JOURNAL_ENTRY.findall(text))
    formula_count = len(RE_FORMULA.findall(text))
    tax_count = len(RE_TAX_RATE.findall(text))
    choice_count = len(RE_CHOICE_Q.findall(text))

    if journal_count >= 3:
        types.add("journal_entries")
    if formula_count >= 3:
        types.add("formulas")
    if tax_count >= 2:
        types.add("tax_rules")
    if choice_count >= 2:
        types.add("practice_problems")

    # If no strong signal, fall back to hint
    if not types:
        types.add(hint)

    return sorted(types)


def parse_headings(text: str) -> list[str]:
    """Extract heading lines from text."""
    headings = []
    for match in RE_HEADING.finditer(text):
        # Get the full line containing the match
        start = text.rfind("\n", 0, match.start()) + 1
        end = text.find("\n", match.end())
        if end == -1:
            end = len(text)
        line = text[start:end].strip()
        if 3 < len(line) < 120:  # reasonable heading length
            headings.append(line)
    return headings


def extract_journal_entries(text: str) -> list[dict]:
    """Extract structured journal entries (debit/credit pairs) from text."""
    entries = []
    # Split by common entry separators
    lines = text.split("\n")
    current_entry = []
    current_title = ""

    for line in lines:
        stripped = line.strip()
        if not stripped:
            if current_entry:
                entries.append({
                    "title": current_title,
                    "lines": current_entry[:],
                })
                current_entry = []
                current_title = ""
            continue

        if RE_JOURNAL_ENTRY.search(stripped):
            current_entry.append(stripped)
        elif current_entry:
            # Non-journal line after journal lines = continuation or note
            current_entry.append(stripped)
        else:
            # Potential title/context before journal entry
            current_title = stripped

    if current_entry:
        entries.append({"title": current_title, "lines": current_entry})

    return entries


def extract_formulas(text: str) -> list[str]:
    """Extract lines containing formulas."""
    formulas = []
    for line in text.split("\n"):
        stripped = line.strip()
        if RE_FORMULA.search(stripped) and len(stripped) > 8:
            formulas.append(stripped)
    return formulas


def process_pdf(pdf_path: Path, label: str, subject: str, hint: str) -> dict:
    """Process a single PDF into a structured extraction result."""
    pages = extract_pdf_text(pdf_path)
    if not pages:
        return None

    all_headings = []
    all_journal_entries = []
    all_formulas = []
    all_content_types = set()
    full_text_pages = []

    for pg in pages:
        text = pg["text"]
        page_num = pg["page"]

        # Classify
        types = classify_page(text, hint)
        all_content_types.update(types)

        # Extract structured elements
        headings = parse_headings(text)
        all_headings.extend(headings)

        if "journal_entries" in types:
            je = extract_journal_entries(text)
            for entry in je:
                entry["page"] = page_num
            all_journal_entries.extend(je)

        if "formulas" in types:
            fmls = extract_formulas(text)
            all_formulas.extend(fmls)

        # Store page text for full-text search
        full_text_pages.append({
            "page": page_num,
            "text": text,
            "content_types": types,
        })

    # Deduplicate formulas
    all_formulas = list(dict.fromkeys(all_formulas))

    return {
        "source_file": str(pdf_path.relative_to(PROJECT_ROOT)),
        "file_hash": file_hash(pdf_path),
        "label": label,
        "subject": subject,
        "content_types": sorted(all_content_types),
        "page_count": len(pages),
        "total_chars": sum(len(pg["text"]) for pg in pages),
        "headings": all_headings[:200],  # cap for sanity
        "journal_entries": all_journal_entries,
        "formulas": all_formulas,
        "pages": full_text_pages,
    }


def process_xlsx(xlsx_path: Path, label: str, subject: str, hint: str) -> dict:
    """Process a single XLSX into a structured extraction result."""
    sheets = extract_xlsx_text(xlsx_path)
    if not sheets:
        return None

    all_rows = []
    for s in sheets:
        all_rows.extend(s["rows"])

    # Detect journal entries in XLSX rows
    journal_entries = []
    for row_text in all_rows:
        if RE_JOURNAL_ENTRY.search(row_text):
            journal_entries.append({"title": "", "lines": [row_text]})

    return {
        "source_file": str(xlsx_path.relative_to(PROJECT_ROOT)),
        "file_hash": file_hash(xlsx_path),
        "label": label,
        "subject": subject,
        "content_types": [hint],
        "sheet_count": len(sheets),
        "total_rows": len(all_rows),
        "journal_entries": journal_entries,
        "sheets": sheets,
    }


# ── Main ─────────────────────────────────────────────────────────────────


def build_file_list() -> list[dict]:
    """Build the complete list of files to process."""
    files = []

    # Priority individual files
    for entry in PRIORITY_FILES:
        path = DOC_BASE / entry["rel_path"]
        if path.exists():
            files.append({
                "path": path,
                "label": entry["label"],
                "subject": entry["subject"],
                "hint": entry["content_hint"],
                "format": "pdf",
            })
        else:
            print(f"  WARN: priority file not found: {entry['rel_path']}")

    # Batch directories
    for batch in BATCH_DIRS:
        dir_path = DOC_BASE / batch["rel_dir"]
        if not dir_path.is_dir():
            print(f"  WARN: batch dir not found: {batch['rel_dir']}")
            continue
        for pdf in sorted(dir_path.glob("*.pdf")):
            # Generate label from filename
            safe_name = re.sub(r"[^\w\-]", "_", pdf.stem)[:60]
            is_answer = "答案" in pdf.name
            files.append({
                "path": pdf,
                "label": f"{batch['label_prefix']}_{safe_name}",
                "subject": batch["subject"],
                "hint": "answers" if is_answer else batch["content_hint"],
                "format": "pdf",
            })

    # XLSX files
    for entry in XLSX_FILES:
        path = DOC_BASE / entry["rel_path"]
        if path.exists():
            files.append({
                "path": path,
                "label": entry["label"],
                "subject": entry["subject"],
                "hint": entry["content_hint"],
                "format": "xlsx",
            })

    return files


def main():
    parser = argparse.ArgumentParser(description="Extract CPA study materials")
    parser.add_argument("--dry-run", action="store_true", help="List files without extracting")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    files = build_file_list()

    print(f"Found {len(files)} files to process")
    if args.dry_run:
        for f in files:
            print(f"  [{f['format']}] {f['label']}: {f['path'].name}")
        return

    manifest_entries = []
    stats = {"total": 0, "success": 0, "failed": 0, "skipped": 0}
    stats_by_type = {}

    for i, f in enumerate(files, 1):
        path = f["path"]
        label = f["label"]
        fmt = f["format"]
        print(f"[{i}/{len(files)}] {fmt.upper()} {label}: {path.name}")
        stats["total"] += 1

        if fmt == "pdf":
            result = process_pdf(path, label, f["subject"], f["hint"])
        elif fmt == "xlsx":
            result = process_xlsx(path, label, f["subject"], f["hint"])
        else:
            print(f"  WARN: unsupported format {fmt}")
            stats["skipped"] += 1
            continue

        if result is None:
            print(f"  WARN: no content extracted")
            stats["failed"] += 1
            continue

        # Write individual JSON
        out_file = OUTPUT_DIR / f"{label}.json"
        with open(out_file, "w", encoding="utf-8") as fh:
            json.dump(result, fh, ensure_ascii=False, indent=2)

        # Track stats
        stats["success"] += 1
        for ct in result.get("content_types", []):
            stats_by_type[ct] = stats_by_type.get(ct, 0) + 1

        je_count = len(result.get("journal_entries", []))
        fm_count = len(result.get("formulas", []))
        pg_count = result.get("page_count", result.get("sheet_count", 0))
        chars = result.get("total_chars", result.get("total_rows", 0))

        manifest_entries.append({
            "label": label,
            "source_file": result["source_file"],
            "file_hash": result["file_hash"],
            "subject": result.get("subject", ""),
            "content_types": result.get("content_types", []),
            "pages_or_sheets": pg_count,
            "total_chars_or_rows": chars,
            "journal_entries_count": je_count,
            "formulas_count": fm_count,
            "output_file": f"data/extracted/cpa/{label}.json",
        })

        print(f"  OK: {pg_count} pages, {chars} chars, {je_count} entries, {fm_count} formulas -> {out_file.name}")

    # Write manifest
    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generator": "src/extract_cpa_materials.py",
        "source_dir": str(DOC_BASE.relative_to(PROJECT_ROOT)),
        "output_dir": str(OUTPUT_DIR.relative_to(PROJECT_ROOT)),
        "stats": {
            **stats,
            "by_content_type": stats_by_type,
        },
        "files": manifest_entries,
    }
    manifest_path = OUTPUT_DIR / "manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"Extraction complete: {stats['success']}/{stats['total']} files")
    print(f"Content types: {stats_by_type}")
    print(f"Manifest: {manifest_path}")


if __name__ == "__main__":
    main()
